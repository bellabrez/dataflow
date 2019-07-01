from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Fill
import json
import os
import bigbadbrain as bbb

def load_json(file):
    with open(file, 'r') as f:
        data = json.load(f)
    return data

def main():
    filename='/oak/stanford/groups/trc/data/Brezovec/2P_Imaging/20190101_walking_dataset/master_2P.xlsx'
    wb = load_workbook(filename=filename, read_only=False)
    ws = wb.active
    #ws = wb['big_data']
    #wb = Workbook()
    #wb = wb.load()
    # grab the active worksheet
    #col = ws.column_dimensions['B']
    #col.font = Font(bold=True)
    # Data can be assigned directly to cells
    #ws['A1'] = 42

    root_directory = '/oak/stanford/groups/trc/data/Brezovec/2P_Imaging/20190101_walking_dataset/'
    fly_folders = [os.path.join(root_directory,x) for x in os.listdir(root_directory) if 'fly' in x]
    bbb.sort_nicely(fly_folders)
    #fly_folder = '/oak/stanford/groups/trc/data/Brezovec/2P_Imaging/20190101_walking_dataset/fly_36'
    for fly_folder in fly_folders:
        print(fly_folder)
        # If no fly.json, just skip (atleast one fly in empty for some reason)
        try:
            fly_file = os.path.join(fly_folder, 'fly.json')
            fly_data = load_json(fly_file)
        except:
            continue

        expt_folders = []
        expt_folders = [os.path.join(fly_folder,x) for x in os.listdir(fly_folder) if 'func' in x]
        bbb.sort_nicely(expt_folders)
        for expt_folder in expt_folders:

            expt_file = os.path.join(expt_folder, 'expt.json')
            expt_data = load_json(expt_file)

            # Occasionally a fly may not have an imaging folder (if only fictrac was recorded for example)
            try:
                scan_file = os.path.join(expt_folder, 'imaging', 'scan.json')
                scan_data = load_json(scan_file)
                print(type(scan_data['x_voxel_size']))
                scan_data['x_voxel_size'] = '{:.1f}'.format(scan_data['x_voxel_size'])
                scan_data['y_voxel_size'] = '{:.1f}'.format(scan_data['y_voxel_size'])
                scan_data['z_voxel_size'] = '{:.1f}'.format(scan_data['z_voxel_size'])
                print(type(scan_data['x_voxel_size']))
            except:
                scan_data['laser_power'] = None
                scan_data['PMT_green'] = None
                scan_data['PMT_red'] = None
                scan_data['x_dim'] = None
                scan_data['y_dim'] = None
                scan_data['z_dim'] = None
                scan_data['x_voxel_size'] = None
                scan_data['y_voxel_size'] = None
                scan_data['z_voxel_size'] = None

            visual_file = os.path.join(expt_folder, 'visual', 'visual.json')
            try:
                visual_data = load_json(visual_file)
                visual_input = visual_data[0]['name'] + ' ({})'.format(len(visual_data))
            except:
                visual_input = None

            # Get fly_id
            fly_folder = os.path.split(os.path.split(expt_folder)[0])[-1]
            fly_id = fly_folder.split('_')[-1]

            # Get expt_id
            expt_id = expt_folder.split('_')[-1]

            # Append the new row
            new_row = []
            new_row = [int(fly_id),
                       int(expt_id),
                       fly_data['date'],
                       expt_data['brain_area'],
                       fly_data['genotype'],
                       visual_input,
                       fly_data['notes'],
                       expt_data['notes'],
                       expt_data['time'],
                       fly_data['circadian_on'],
                       fly_data['circadian_off'],
                       fly_data['gender'],
                       fly_data['age'],
                       fly_data['temp'],
                       scan_data['laser_power'],
                       scan_data['PMT_green'],
                       scan_data['PMT_red'],
                       scan_data['x_dim'],
                       scan_data['y_dim'],
                       scan_data['z_dim'],
                       scan_data['x_voxel_size'],
                       scan_data['y_voxel_size'],
                       scan_data['z_voxel_size']]

            #if visual_data is not None:
            #    new_row['visual_stimuli'] = visual_data
            #if fictrac_data is not None:
            #    new_row['fictrac'] = fictrac

            ws.append(new_row)

    # Save the file
    filename='/oak/stanford/groups/trc/data/Brezovec/2P_Imaging/20190101_walking_dataset/master_2P_new.xlsx'
    wb.save(filename)

if __name__ == '__main__':
    main()
